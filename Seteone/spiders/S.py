import datetime
# import mysqldb
from requests.auth import HTTPDigestAuth
from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.orm import declarative_base, sessionmaker
import json
import time
import requests
import scrapy
from openpyxl import Workbook
wb = Workbook()
from openpyxl import load_workbook
import pandas as pd
from scrapy import Request
from ..models import Column#, db_connect #, create_items_table#, Character

Base = declarative_base()
#
#
def db_connect():
    username = 'avnadmin'
    password = 'AVNS_OZaFMf--rCU9QzzfxpE'
    host = 'mysql-198e88e7-sszapis-39be.g.aivencloud.com'
    port = '22057'  # Порт по умолчанию для MySQL
    database = 'defaultdb'
    DATABASE_URL = f"mysql+mysqlconnector://{username}:{password}@{host}:{port}/{database}"  # # "mysql+mysqldb://{username}:{password}@{hostname}/{databasename}" mysqldb
    # engine = create_engine(DATABASE_URL)
    return create_engine(DATABASE_URL, echo=True)

def create_items_table(engine):
    Base.metadata.create_all(engine)
class Character(Base):
    __tablename__ = 'products'
    Shop_Number = Column(Integer, primary_key=True)
    Title = Column(String(255))
    Date = Column(String(255))
    _40202_5G = Column(String(255))
    # _402030G = Column(String(255))
    # _30102_5G = Column(String(255))
    # _30132_5G = Column(String(255))
    # _50102_5G = Column(String(255))
    # _301232OZC = Column(String(255))
    # _50101G = Column(String(255))
    # _301032OZ = Column(String(255))
    # _402032OZRTUN = Column(String(255))
    # _30932OZ = Column(String(255))
    # _10750444 = Column(String(255))
    # _3092_5G = Column(String(255))
    # _3081G = Column(String(255))
    # _301330G = Column(String(255))
    # _3012250GC = Column(String(255))
    # _108SS = Column(String(255))
    # _30832OZ = Column(String(255))
    # _30732OZ_NLA = Column(String(255))
    # _301332OZ = Column(String(255))
    # _30830G = Column(String(255))
    # _30130G = Column(String(255))
    # _308250G = Column(String(255))
    # _501030G = Column(String(255))
    # _101B = Column(String(255))
    # _1017PEACH = Column(String(255))
    # _501055G = Column(String(255))
    # _10140CA = Column(String(255))
    # _10140 = Column(String(255))
    # _10940CPG = Column(String(255))
    # _109SSCPG = Column(String(255))

class SSpider(scrapy.Spider):
    name = "S"

    def start_requests(self):
        headers = {
          'accept': '*/*',
          'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
          'cache-control': 'no-cache',
          # 'cookie': 'ROUTE=.accstorefront-5944bff4d7-mtq99; accountPageId=; allowBlockAccess=none; siteone-cart=f94c3713-3ce5-4d8c-baa8-97b70641aee2; gls=1003; AMCVS_44BC3F5859E47A930A495D8E%40AdobeOrg=1; _gcl_au=1.1.334648884.1722361175; _lfa=LF1.1.6ab2d15a92daf043.1722361176041; com.silverpop.iMAWebCookie=d1d02d16-68ab-cbe2-85cd-ebb2a336e139; s_cc=true; _gid=GA1.2.1174516156.1722361257; _fbp=fb.1.1722361259868.891558262566629955; JSESSIONID=4C318D3CB06F5E060D9C59E0471181ED.accstorefront-5944bff4d7-mtq99; _abck=8207946FB29DCFE5B21703E38A85B0EA~0~YAAQWHjOFw/pgAaRAQAAyd90CQwTKKUmKhh7yVcpn68T6EjFq05i5TsDLqZWoDRPVykAycWq53OnqhHEO4jXv3N0Z0ncg1IpznQ+FKhRp8hPonue/ZrEiAHYfbaFGjiy/1T6jmNjbnGtVlzLg/p1VB6gGhUkasuJ1csTaLurfM2Z8vJffXgsjelEoXgTycFacjBq+Ka9AfeTfLy66Z1MtFrTT4OnnnJwg/nZiZszP7KXkPH78pmvF7ygk5MRF9jj59rsaNpO+FEnKYZ/9iLLKSuTu2/Sl7UG/2zsYBjkNKirj8UNU260M+vVFxyoE0U9EqNTVgjFyOoeMwnhlQfRfJJHiAFzGHT4WwXRtStTqMp+LKm9bo/iNK292F8Zt87lQ36gH4UKAfy8Vp6QCw0X3848q52cKy2Jew==~-1~-1~-1; s_lv_s=Less%20than%201%20day; _ga_NB436M2T0P=deleted; com.silverpop.iMA.session=96b9332c-8ad3-c1a8-bd69-1e35b4156c25; AMCV_44BC3F5859E47A930A495D8E%40AdobeOrg=1176715910%7CMCIDTS%7C19935%7CMCMID%7C86505869982574825551914714817971891429%7CMCAAMLH-1723052924%7C7%7CMCAAMB-1723052924%7CRKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y%7CMCOPTOUT-1722455324s%7CNONE%7CMCAID%7C331A514E3CF025EB-60001A9AA07F23C8%7CvVersion%7C5.4.0; bm_mi=014DC46BABCF1534D059BB9AB92C3752~YAAQUnjOF5JrkuuQAQAA/znvCRhqXEf0cXQTcj8gmTAAjpXGZjcbrLSi9n2WIEaSMthN4mog6bkiOy6wCED55gdxOpVyQzv6oQ5x3vpQa7pyDz9npE6lEPqb1f3e+TdiTu2UXtmEFhZ3aZYG0/KF2GaIZgk6lL2GNyTudJoeDfDa2H94cTwEuzJRxStILgx+nIZqUKlXaI9r8LUoCjGcE+qJyFukQNLZuIdzkCfKIp0hGHhYWMuoL0Cnbmi29PsLQ+GLR2CUikit/hWC49pnhcnoKZBDO8yJ5v631FS1uU/ZtSel78yPh38bT3IDA6IZGaTssF2LpfJrxEB1xg==~1; ak_bmsc=5F417C344ED85E9014178EF2623E7632~000000000000000000000000000000~YAAQUnjOFxyKkuuQAQAAIHvvCRgXjLpZvdpynxO5N4Dsis5y3c/BIHdcSBsbkhbOIeJgVD1yZD8yoJwpgK3XO4vTzvfbMeu/Bkwhlf5+5nkL+d3TdxcYxel2GbAH1VdNzh4qtp4jmznx82LvrzPlseCXzwHEJKSSeT/8qwqSfYYVrIfJSDr1t90DeJDhWGDU4OMIybn21y6qAxPxNA8J06IB5cRUx3kn7hWBXNN/vxrztu7SnCkW5ZDqYYHX4J4Vm2mxCWsB+kxTZYkg9zKS6K40o6l9yxHBc3mHl5vaHsUMO91Zjr/nZkkRz//+udtrjwSa8kD0F/+H7oVFJZNZeX9wDitLpmNT11pbxvQr3yDp8cr+qTBt7mu2kjn8TWoOYeRPedGqR397fl1JL2mES/hXV2hrThD9tGw70BOasdCu2a47i/45uFWAHtmNFLD58ejCIPYmEYuIFc2blfFhVCV83v0siHaEc9geJro9AgoplZHU724g30nNtBE9iqj5; csc=1002; com.silverpop.iMA.page_visit=-1486763195:941829075:1169047474:-1882466775:931128564:; s_sq=%5B%5BB%5D%5D; bm_sv=055332EDC5CA507351B81E609B1E251A~YAAQZ1XIF+tFqeOQAQAAjfvzCRjgmB9Qc8U63AMXjptTmZKZN7pAnzEYTSAS2uSUiaO93GZQclfUC/ConyCW51juyAoEdpLwgmzCXpcK8NUdhZEpsKmSM7LqWclWJM6Yrsc1AqX/H4cPk/DXTSfDBtogc8VrErPbgQTuDi64bBOkllBFWiu5t0Lhq7GM1mJrbKirz+w1bFxV/NLUGheLj8lmi5Lq0yyt2AkOS96Uf0QbcBpoR9WGpwELzujCZTkLHPA=~1; bm_sz=06B48E3AFD3D9B0CE9B9BE6AA425B375~YAAQZ1XIF+xFqeOQAQAAjfvzCRjuxPxjuQ48kh1ZwxHmqMOzkOE4Jqi8Us+XdqGyVIpdBUbfgkZBoZvbZYdIOxNhwzZ/27qsFAG0yXNh1o8FMyjxsm1MEq/EDLxVAZGrQoNoODll599PSDM0r6YjpOMVIfBz0JjIjD3fvrAeYT+BmCAv39c3YQyUjfGC8MuTEwyc7DkymKsYBkvAqXLRkGd42X46IQV0ufdWqLb9hKZCF+GaiVz6rtIvxaR0MiYE207vumKtQ73kmTvgUbpHhWV962UlDTfDJgut0Zo1plYk7NZ335t3W8Ljjfgv++ry+WBcz5tfDc8XNU7KAqzZOOWK6oXS+lqhcz+4eo7RAFpgOfy5yan2AaaqH+OsqY0++WutBINJYE23TdiQShWxsuoFkY/7V0vSh9QCB2MafRg+J0ZQ0vT4TwUUHKUONXp7eZVB4K6YDvrdDkWwxwSM31C6fvZU47BPhF4W8ZJIcCfOuvDeduvtRalUQIX71YVy/BLVA/nKTMDC25v/PDHDzyi0Pd19nkuh3VNvJ9ggPyE2~3355440~3486534; gpv=store%3A%20store%20finder; RT="z=1&dm=siteone.com&si=cd40yboksui&ss=lz9lg06e&sl=0&tt=0"; fs_lua=1.1722448880381; fs_uid=#HHD7H#1d29774c-bbbb-4038-81af-3e66b4baa222:3bb995e5-e976-432e-9837-b617e8926329:1722444332365::23#/1753897377; _rdt_uuid=1722361173867.dbd3b2ba-9209-40c5-8bd4-b97e611d587e; _gat_gtag_UA_23083629_11=1; _ga_QWN01LTG1X=GS1.1.1722444334.4.1.1722448881.0.0.0; _ga=GA1.1.1272780326.1722361176; _ga_NB436M2T0P=GS1.1.1722444334.4.1.1722448885.46.0.0; s_ppvl=My%2520Account%2520%253A%2520Login%2520Page%2C49%2C49%2C809%2C1731%2C809%2C1730%2C973%2C1.11%2CP; s_ppv=store%253A%2520store%2520finder%2C60%2C60%2C809%2C1042%2C809%2C1730%2C973%2C1.11%2CP; s_nr=1722448893117-Repeat; s_lv=1722448893119; _abck=8207946FB29DCFE5B21703E38A85B0EA~-1~YAAQ00A2F/PF3OKQAQAAP7r8CQwylKXdrqPpo4W7SWaJyBChVljDWkbJcagkwBw4VjkIWlgAZIZJS7Sj/dNd6dqCRDFOicIjBFle3KcrRaKI+1xu87chmVyjjxwCWiIGc+NKotXL7MJ4gSdTvGDypUo1ecO0Ld1KOJ/fAYLGOMWgdsEdTSVZBE1A/cfSSEsCpHBWzGDfdHiG1vAvS7w2gEQPu1/cKa0mGZgcc9b/NrOqiWh5xwHBy+hecASL0fsOe92/czabDRO4FHBhF+JgDug3rc2/FVp4CGVGZ1vK4PUjmU3UAT7fFvl2AJDqDYWlTSRdyt7HR5Hc42aBiDFeQfIL1mo/S2YuFAw1bfRmT53YVlCaRZGkRdTCaUDxm5dvMU7ebpFScrIHSFy3EKeq7bZ566m4bBYaMw==~0~-1~-1; bm_mi=014DC46BABCF1534D059BB9AB92C3752~YAAQUnjOF8ExkOuQAQAAk0/rCRhyrf8Gy2jEIxMlwwaiNwNtsi+CGox5GC8N7ufRHwTwbDcrd6lR5t/jba3s0ZHydP8d5haPCx8FL3gduFnuvowcKAe5L4CwXGF5KIyoZpUkrzhypwt9wXdDNf8r6UsdIGBrNMT5yGOPfauEGu1BgtlPv2+Gw0q2x4NW3zRrdwv6PSOKT7/HGEAP3bouvraVyctuSizJt0Yx7hWmkw5uv26/Lv/HmR6418RBUSa6FA3BcnUT8+RbTDc0jhQKV1RwF1ul3UTWtGz+fzMh9R7fIa9zor5NUB0cjyypAwZT+X3ohd8oNx4ByalTRMoig2Gwl7g=~1; bm_sv=055332EDC5CA507351B81E609B1E251A~YAAQ00A2F/TF3OKQAQAAP7r8CRioDiTMiL4u8nRNd1kYlRrCaR1FtN3sfLksIu2s6da9/aeApM5okgpCqGcH9k880uw9O5yg9q0+inxJ9VaVx5qPK2zFloUSCc1iGFCz/8nfRBV8AoyEOoKzR8qQmwJym4CwNOln792codcy35muuz7dBomi2Hgr+t+uWWWF7Ci6evpoTI2inDOTFt81RutNg2KqtZlBJ8+QxX/T+NfFbY2ehmyQGyU0REzxw6sf8rs=~1; JSESSIONID=4C318D3CB06F5E060D9C59E0471181ED.accstorefront-5944bff4d7-mtq99; ROUTE=.accstorefront-5944bff4d7-h2hj2; csc=485; siteone-cart=f94c3713-3ce5-4d8c-baa8-97b70641aee2',
          'pragma': 'no-cache',
          'priority': 'u=1, i',
          'referer': 'https://www.siteone.com/en/store-finder',
          'sec-ch-ua': '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
          'sec-ch-ua-mobile': '?0',
          'sec-ch-ua-platform': '"Windows"',
          'sec-fetch-dest': 'empty',
          'sec-fetch-mode': 'cors',
          'sec-fetch-site': 'same-origin',
          'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36',
          'x-requested-with': 'XMLHttpRequest'
        }
        # wb = load_workbook(r'/home/Reyter/Siteone/Siteone/SiteOne - Mirimichi Green Products.xlsx')
        wb = load_workbook(r'D:\PycharmProjects\Upwork\Afonso T\siteone.com\Site\Site\SiteOne - Mirimichi Green Products.xlsx')
        ws = wb['Sheet1']
        dict_list = []
        dict_list2 = []
        date = str(datetime.datetime.now()).split('.')[0]
        for rn in range(2, 3, 1):  # 32
            all_names = []
            each_sku = str(ws[f'B{rn}'].value)
            each_code = str(ws[f'C{rn}'].value).split('/')[-1].replace(' ','')
            coordinates = ['33.4443871&longitude=-86.8424924', '33.4951698&longitude=-112.0273767','34.862214&longitude=-92.3973805','33.862849&longitude=-117.8426362','40.4823&longitude=-104.89214',
                           '41.7736739&longitude=-72.652851','33.8540667&longitude=-78.8269606','27.9732131&longitude=-82.3393785','33.934587&longitude=-84.503936','43.6068927&longitude=-116.3599583','39.747609&longitude=-89.712525',  '39.9092412&longitude=-86.2582754',
                           '41.6613842&longitude=-93.77193','39.2153061&longitude=-94.4810336','38.0800616&longitude=-84.506935','30.3076716&longitude=-89.8920859','45.413574&longitude=-122.7503403', '39.9092412&longitude=-86.2582754',
                           '39.1790071&longitude=-76.7629581','42.1842874&longitude=-83.2543515', '44.9921294&longitude=-93.4077615', '39.747609&longitude=-89.712525', '40.9361299&longitude=-98.3861278', '36.1042748&longitude=-115.1941607', '43.1863094&longitude=-71.4870034',
                           '39.4093662&longitude=-74.5373327', '35.1729342&longitude=-106.6045778','40.9324197&longitude=-73.7643783', '35.0443725&longitude=-78.8285755','29.1392631&longitude=-82.2190105', '39.9882702&longitude=-83.1383765', '35.6192275&longitude=-97.5055128',
                           '44.0465755&longitude=-123.1542307', '40.4549389&longitude=-80.1068607', '41.7406048&longitude=-71.4401636', '34.0104827&longitude=-117.9569816', '32.7976093&longitude=-80.1277233',  '41.9473821&longitude=-88.0206873', '36.1175391&longitude=-86.7522436', '32.8044472&longitude=-96.6660467',
                           '40.5732892&longitude=-112.035574', '37.5990991&longitude=-77.5100597', '47.5822583&longitude=-122.164614', '43.210205&longitude=-87.9863891']
            for b in coordinates:
                each_url = f'https://www.siteone.com/en/store-finder/stores?miles=1000&latitude={b}&productCode={each_code}'
                # time.sleep(3)
                r = requests.get(each_url, headers=headers)
                data = json.loads(r.text)
                data2 = data["nearBy"]['data']
                for i in data2:
                    try:
                        title = i['name']
                    except:
                        title = ''
                    try:
                        stock = i["stockDetail"]
                    except:
                        stock = ''
                    if title not in all_names:
                        all_names.append(title)
                        shop_code = title.split('#')[1].replace(' ','')
                        compl_data = {
                            'Shop_Number': shop_code,  # []
                            'Title': title,
                            f'{each_sku}': stock
                        }
                        if rn == 2:
                            dict_list.append(compl_data)

                        if rn != 2:
                            compl_data2 = {
                                'Shop_Number': shop_code,
                                'SKU1': each_sku,
                                'SKU': str(each_sku + " " + stock),
                                'SKU2': stock
                            }
                            dict_list2.append(compl_data2)
        engine = db_connect()
        Base.metadata.drop_all(engine)
        # create_items_table(engine)
        Base.metadata.create_all(engine)
        Session = sessionmaker(bind=engine)
        session = Session()
        for d2 in dict_list2:
            for d1 in dict_list:
                if d2['Shop_Number'] == d1['Shop_Number']:
                    try:
                        d1[d2['SKU1']] = d2['SKU2']
                    except:
                        pass
        for data_insert in dict_list:
            try:
                sh_num = data_insert['Shop_Number']
            except:
                sh_num = ''
            try:
                title= data_insert['Title']
            except:
                title = ''
            try:
                insertion_query = Character(Shop_Number= sh_num, Title = title,Date= date, _40202_5G= data_insert['40202.5G'])#, _402030G= data_insert['402030G'], _30102_5G = data_insert['30102.5G'],
                    # _30132_5G = data_insert['30132.5G'], _50102_5G = data_insert['50102.5G'], _301232OZC= data_insert['301232OZC'],_50101G = data_insert['50101G'],_301032OZ= data_insert['301032OZ'],
                    # _402032OZRTUN = data_insert['402032OZRTUN'], _30932OZ = data_insert['30932OZ'], _10750444 =data_insert['10750444'], _3092_5G = data_insert['3092.5G'], _3081G = data_insert['3081G'],
                    # _301330G = data_insert['301330G'], _3012250GC = data_insert['3012250GC'], _108SS = data_insert['108SS'], _30832OZ = data_insert['30832OZ'], _30732OZ_NLA = data_insert['30732OZ-NLA'],
                    # _301332OZ = data_insert['301332OZ'], _30830G = data_insert['30830G'], _30130G = data_insert['30130G'], _308250G = data_insert['308250G'], _501030G = data_insert['501030G'], _101B = data_insert['101B'],
                    # _1017PEACH = data_insert['1017PEACH'], _501055G = data_insert['501055G'], _10140CA = data_insert['10140CA'], _10140 = data_insert['10140'], _10940CPG = data_insert['10940CPG'],
                    # _109SSCPG = data_insert['109SSCPG'])
                session.add(insertion_query)
            except:
                pass
        try:
            session.commit()
            session.close()
        except:
            pass




